# -*- coding: utf-8 -*-
"""
Created on Mon Feb 29 19:32:13 2016

@author: www
"""


from openpyxl import load_workbook

def parsefile(filename, acadyear, withhead=False):
    wb=load_workbook(filename, data_only=True)
    results=[]
    wb.get_sheet_names()
    ws=wb.get_sheet_by_name('OVERALL Results')
    modulename=ws['A1'].value[:7]
    modulelevel=modulename[2]
    modulesem=modulename[3]
    row=4
    entry=list(ws['A%s:K%s'%(row,row)])[0]
    spr=entry[0].value.split('/')[0]
    route=entry[3].value
    d1stat=entry[4].value
    RavB=entry[7].value
    CwkAv=entry[9].value
    d2stat=entry[10].value
    if withhead:
        results.append("\t".join(['Module', 'Level', 'Semester', 'Academic year',spr,route,str(RavB),str(CwkAv),d1stat,d2stat])+'\n')    
    row=5
    entry=list(ws['A%s:K%s'%(row,row)])[0]
    
    while entry[0].value :
        spr=entry[0].value.split('/')[0]
        route=entry[3].value
        d1stat=entry[4].value
        RavB=entry[7].value
        CwkAv=entry[9].value
        d2stat=entry[10].value
        row+=1
        results.append("\t".join([modulename, modulelevel, modulesem, acadyear,spr,route,str(RavB)[:5],str(CwkAv)[:5],d1stat,d2stat])+'\n')    
        entry=list(ws['A%s:K%s'%(row,row)])[0]
    return results
    
fh=open('moduleresults.txt')
ofh=open('output.txt','w')    
flag=True
for f in fh.readlines():
    acadyear=f.split('/')[0]
    filename="/l/%s"%f
    res=parsefile(filename, acadyear, withhead=flag)
    flag=False
    ofh.write(''.join(res))
ofh.close()